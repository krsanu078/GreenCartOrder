#-------------------------importing libraries-----------------------------
import time
import openpyxl

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

from selenium.webdriver import ActionChains
from selenium.webdriver.support.select import Select

#---------------------------------------------------------------------------
#-----------------Giving inputs using Excel sheet---------------------------   
data_input=openpyxl.load_workbook(r"C:\Users\skuma183\Pictures\GreenCart Order System\GreenCartOrder.xlsx")        
sheet = data_input.active
vegetables=[]
quantity_in_kg=[]
for i in range(2,sheet.max_row+1):
    vegetables.append(sheet.cell(row=i, column=1).value)
    quantity_in_kg.append(int(sheet.cell(row=i, column=2).value))

#----------------------------------------------------------------------------
# ---------------Browser Invocation------------------------------------------
driver = webdriver.Chrome(executable_path=r"C:\Users\skuma183\Pictures\chromedriver_win32.exe")
driver.maximize_window()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/seleniumPractise/")

#------------------------------------------------------------------------------
#-------------------Searching and Adding Vegetables to cart---------------------------
for i in range(len(vegetables)):
    vegetable=vegetables[i]
    driver.find_element(By.CSS_SELECTOR, "input.search-keyword").send_keys(vegetable)
    time.sleep(4)
    driver.find_element(By.XPATH,'//input[@class="quantity"]').clear()

    quantity=quantity_in_kg[i]    
    driver.find_element(By.XPATH,'//input[@class="quantity"]').send_keys(quantity)
    driver.find_element(By.XPATH,'//button[text()="ADD TO CART"]').click()
    driver.find_element(By.CSS_SELECTOR, "input.search-keyword").clear()

# -------------------------------------------------------------------------------
#------------Proceed to checkout and order verification-----------------------------
list = []
list2 = []
driver.find_element(By.CSS_SELECTOR, "img[alt='Cart']").click()
Product=driver.find_elements(By.XPATH,'//p[@class="product-name"]')
for i in range (sheet.max_row-1):
    list.append(Product[i].text)


driver.find_element(By.XPATH, "//button[text()='PROCEED TO CHECKOUT']").click()
wait = WebDriverWait(driver, 5)
wait.until(expected_conditions.presence_of_element_located((By.CLASS_NAME, "promoCode")))
veggies =driver.find_elements(By.CSS_SELECTOR, "p.product-name")
for l in veggies:
    list2.append(l.text)


assert list == list2

#-------------------------------------------------------------------------------
#-------------------Applying Promocode-------------------------------------------
Promocode=(sheet.cell(row=2, column=3).value).lower()
driver.find_element(By.CLASS_NAME, "promoCode").send_keys(Promocode)
driver.find_element(By.CSS_SELECTOR, ".promoBtn").click()
amount = int(driver.find_element(By.XPATH, '//span[@class="totAmt"]').text)
veggiesAmount = driver.find_elements(By.XPATH, "//tr/td[5]/p")
sum = 0
for v in veggiesAmount:
    sum = sum + int(v.text)

assert sum==amount

msg=(driver.find_element(By.CSS_SELECTOR, "span.promoInfo").text)

if Promocode=="rahulshettyacademy":
    assert msg=="Code applied ..!"
elif Promocode=="":
    assert msg=="Empty code ..!"
else:
    assert msg=="Invalid code ..!"

#----------------------------------------------------------------------------
#-------------------Proceed further to place the order----------------------------------
action=ActionChains(driver)
action.move_to_element(driver.find_element(By.XPATH, '//button[text()="Place Order"]')).perform()
action.click(driver.find_element(By.XPATH, '//button[text()="Place Order"]')).perform()

dropdown = Select(driver.find_element(By.CSS_SELECTOR,"select"))
Country=(sheet.cell(row=2, column=4).value)
dropdown.select_by_value(Country)

driver.find_element(By.XPATH, '//input[@type="checkbox"]').click()
driver.find_element(By.XPATH, '//button[text()="Proceed"]').click()

# ----------------------------------------------------------------------------
#---------------order placed message verification--------------------------------

msg=driver.find_element(By.XPATH,'//span[text()="Thank you, your order has been placed successfully "]').text
final_msg= "placed successfully" in msg


assert final_msg== True
